#!/usr/bin/env python3
import os
import re
from enum import Enum, auto
from itertools import groupby
from itertools import zip_longest
import math
from collections import OrderedDict
from scipy import stats
import numpy as np
from openpyxl import Workbook
import numbers

"""
Directory wherein all experimental data is stored. Can be recursively organized.
"""
LOCALDIR = 'Data/testing'

"""
Bin size for latency frequency distributions.
"""
LATENCYSTEP = 0.1


class Presets(Enum):
    NIGHT_1 = auto()
    NIGHT_2 = auto()
    NIGHT_3 = auto()
    NIGHT_4 = auto()
    CONTRAST = auto()
    SPATIAL = auto()


"""Experiment-specific values"""
TIMEOUTS = {Presets.NIGHT_3: 30, Presets.NIGHT_4: 10, Presets.CONTRAST: 10, Presets.SPATIAL: 10}
CONTRAST_LVLS = {1: 1, 2: 2, 4: 4, 7: 8, 14: 16, 27: 32, 52: 64, 100: 100}


class ImageTypes(Enum):
    REWARD = "REWARD"
    CONTROL = "CONTROL"


class DoorStates(Enum):
    High, Low = auto(), auto()


class PumpStates(Enum):
    On, Off = auto(), auto()


class Activity(Enum):
    Running, Poking = auto(), auto()


class Mouse:

    def __init__(self, cageNum, rotation_intervals, poke_events):
        self._cageNum = cageNum
        self._rotation_intervals = rotation_intervals
        self._poke_events = poke_events

    @property
    def rotation_intervals(self):
        return self._rotation_intervals

    @property
    def poke_events(self):
        return self._poke_events


class RotationInterval:
    # contiguous series of wheel spins

    def __init__(self, halfTimes, image):
        self._halfTimes = []
        self._image = image
        self.viable = True
        self._rpms = []  # instantaneous speeds in rotations per minute
        raw_rpms = []  # prone to error, will be refined

        for i in range(1, len(halfTimes) - 1):
            # instantaneous speeds btwn beginning and end of rotation event
            raw_rpms.append(60 / (halfTimes[i + 1] - halfTimes[i - 1]))

        # instantaneous speeds at beginning and end of rotation event
        raw_rpms.insert(0, 30 / (halfTimes[1] - halfTimes[0]))
        raw_rpms.append(30 / (halfTimes[-1] - halfTimes[-2]))

        erratic = []
        for i in range(1, len(halfTimes) - 1):
            if raw_rpms[i] > 200:
                erratic.append(i)
        for i in range(1, len(halfTimes) - 1):
            if i not in erratic:
                self._rpms.append(raw_rpms[i])
                self._halfTimes.append(halfTimes[i])
        if len(self._halfTimes) < 2:
            self.viable = False

    def __hash__(self):
        return self.startTime.__hash__()

    def numRotations(self):
        return len(self._halfTimes) // 2

    @property
    def avgSpeed(self):
        # average speed in RPM
        return self.numRotations() * 60 / (self._halfTimes[-1] - self._halfTimes[0])

    @property
    def speeds(self):
        return self._rpms

    @property
    def startTime(self):
        return self._halfTimes[0]

    @property
    def midTime(self):
        return (self._halfTimes[-1] + self.halfTimes[0]) / 2

    @property
    def halfTimes(self):
        return self._halfTimes

    @property
    def image(self):
        return self._image


class PokeEvent:
    # series of repeated pokes

    def __init__(self, doorStates, doorTimes, pumpTimes, pumpStates, image):
        self._doorStates = doorStates
        self._doorTimes = doorTimes
        self._pumpTimes = pumpTimes
        self._pumpStates = pumpStates
        self._image = image
        self._imageAppearanceTime = image.latestAppearanceTime()
        self._imageAppearance = self._image.appearances.get(self._imageAppearanceTime)
        self._imageAppearance.addPokeEvent(self)
        # add this pokeEvent to the image appearance during which it occured
        s, t = self.successfulPokes()
        if s == 1:
            self.latency = t[0] - self._imageAppearanceTime
            self.pokeTime = t[0]
        else:
            self.latency = None
            self.pokeTime = None

    def isSuccess(self):
        successful = False
        for p in self._pumpStates:
            if p is PumpStates.On:
                successful = True
        return successful

    def successfulPokes(self):
        num = 0
        times = []
        for p, t in zip(self._pumpStates, self._pumpTimes):
            if p is PumpStates.On:
                num += 1
                times.append(t - 0.003)  # Pump is activated 3 ms after poke occurs
        return num, times

    def allPokes(self):
        num = 0
        times = []
        for p, t in zip(self._doorStates, self._doorTimes):
            if p is DoorStates.Low:
                num += 1
                times.append(t)
        return num, times

    def unsuccessfulPokes(self):
        successfulPokes = set(self.successfulPokes()[1])
        allPokes = set(self.allPokes()[1])
        unsuccessful = list(allPokes - successfulPokes)
        unsuccessful.sort()
        return len(unsuccessful), unsuccessful

    def totalPokesNoTimeout(self, grace=30):
        # returns total number of pokes EXCLUDING those that are failed due to image timeout
        # i.e. after pump success
        critical_time = None
        for p, t in zip(self._pumpStates, self._pumpTimes):
            if p is PumpStates.On:
                critical_time = t
        if critical_time is None or self.successfulPokes()[0] > 1:
            return self.allPokes()[0]
        else:
            beforeSuccessful = 0
            for dt in self._doorTimes:
                if dt <= critical_time or dt > critical_time + grace:  # 30 seconds grace period
                    beforeSuccessful += 1
            return int(math.ceil(beforeSuccessful / 2))
            # two door states constitute a full poke
            # ceil necessary because only door opening documented before poke

    def drinkTimes(self):
        drinkStart = 0
        drinkTimes = []
        for p, t in zip(self._pumpStates, self._pumpTimes):
            if p is PumpStates.On:
                drinkStart = t
            else:
                drinkTimes.append(t - drinkStart)
        return drinkTimes

    @property
    def startTime(self):
        return self._doorTimes[0]

    @property
    def image(self):
        return self._image

    @property
    def doorTimes(self):
        return self._doorTimes

    @property
    def doorStates(self):
        return self._doorStates

    @property
    def pumpTimes(self):
        return self._pumpTimes

    @property
    def pumpStates(self):
        return self._pumpStates

    @property
    def imageAppearanceTime(self):
        return self._imageAppearanceTime

    @property
    def imageAppearance(self):
        return self._imageAppearance


class Appearance:

    def __init__(self, image, time, old_img):
        self._time = time
        self._image = image
        self._poke_events = []
        if old_img.appearances:
            old_app_num = old_img.appearances.get(old_img.latestAppearanceTime()).rewardSeqNum
            self._rewardSeqNum = 0 if image.imageType == ImageTypes.CONTROL else old_app_num + 1
        else:
            self._rewardSeqNum = 0 if image.imageType == ImageTypes.CONTROL else 1

    def addPokeEvent(self, poke_event):
        self._poke_events.append(poke_event)

    @property
    def poke_events(self):
        return self._poke_events

    @property
    def time(self):
        return self._time

    @property
    def image(self):
        return self._image

    @property
    def rewardSeqNum(self):
        return self._rewardSeqNum


class Image:
    """
    Log of image appearances by time.
    """
    appearanceLog = OrderedDict()

    def __init__(self, name, imageType):
        self.name = name
        assert isinstance(imageType, ImageTypes), 'use ImageType enum to assign images'
        self.imageType = imageType
        self._appearanceTimes = []
        self._appearances = {}

    def __eq__(self, other):
        return self.name == other.name and self.imageType == other.imageType

    def __hash__(self):
        return hash(self.name)

    def incrementAppearances(self, time, old_img):
        self._appearanceTimes.append(time)
        self._appearances[time] = Image.appearanceLog[time] = Appearance(self, time, old_img)

    @property
    def numAppearances(self):
        return len(self._appearanceTimes)

    @property
    def appearanceTimes(self):
        return self._appearanceTimes

    def latestAppearanceTime(self):
        return self._appearanceTimes[-1]

    @property
    def appearances(self):
        return self._appearances

    @staticmethod
    def imageAtTime(time):
        appearances = list(Image.appearanceLog.keys())
        return Image.appearanceLog[max(filter(lambda k: k < time, appearances))].image


# def cumulativeSuccess(poke_events):
#     outcomes = [int(pe.isSuccess()) for pe in poke_events]
#     print("Successful Poke Events: {0}".format(sum(outcomes)))
#     # times = [pe.startTime for pe in poke_events]
#     cumulative_success = 0
#     total = 0
#     cumulative_probabilities = []
#     for outcome in outcomes:
#         cumulative_success += outcome
#         total += 1
#         cumulative_probabilities.append(cumulative_success / total)
#     plt.plot(cumulative_probabilities, marker='.')
#     plt.ylabel('Cumulative Probability')
#     plt.xlabel('Poke Events')
#     plt.title('Poke Success Rate')
#     plt.show()


def getContrast(image):
    if "negative" in image.name.lower():
        return 0

    try:
        contrastVals = [int(s) for s in image.name if s.isdigit()]
        contrastVal = int(''.join(str(digit) for digit in contrastVals))
    except (IndexError, ValueError):
        contrastVal = 100  # if not level specified, assume 100
    return CONTRAST_LVLS.get(contrastVal, contrastVal)


def pokeLatencies(wb, preset):
    """
    Find latencies and associated statistics image-wise for all poke-events. True latencies represent latencies for
    successful pokes following a reward image appearance, whereas All latencies include missed reward images, using
    the image reset time as a placeholder estimate.
    This function produces 4 excel workbooks per worksheet.
    """

    outputCSV = wb.active
    allLatencies = []
    trueLatencies = []
    rewardTimes = []
    imageWiseTrueLatencies = {}
    imageWiseAllLatencies = {}
    imageWiseTrueLatencies_1st = {}
    imageWiseAllLatencies_1st = {}
    outProxy = [[], ['Time of REWARD', 'Image Contrast Level', 'Latencies (sec)']]

    # contrast instead of name
    # add time in hours
    for ap in Image.appearanceLog.values():
        if ap.image.imageType != ImageTypes.REWARD:
            continue
        # elif ap.rewardSeqNum != 1:  # only first appearances should be considered
        #     continue

        contrastLevel = getContrast(ap.image)

        if not ap.poke_events:
            if imageWiseAllLatencies.get(ap.image) is None:
                imageWiseAllLatencies[ap.image] = []
            if imageWiseAllLatencies_1st.get(ap.image) is None:
                imageWiseAllLatencies_1st[ap.image] = []
            if TIMEOUTS.get(preset) is None:
                continue
            outProxy.append([ap.time, contrastLevel, TIMEOUTS.get(preset)])
            allLatencies.append(TIMEOUTS.get(preset))
            imageWiseAllLatencies[ap.image].append(TIMEOUTS.get(preset))
            rewardTimes.append(ap.time)
            if ap.rewardSeqNum == 1:
                imageWiseAllLatencies_1st[ap.image].append(TIMEOUTS.get(preset))

        else:
            if imageWiseTrueLatencies.get(ap.image) is None:
                imageWiseTrueLatencies[ap.image] = []
            if imageWiseAllLatencies.get(ap.image) is None:
                imageWiseAllLatencies[ap.image] = []
            if imageWiseTrueLatencies_1st.get(ap.image) is None:
                imageWiseTrueLatencies_1st[ap.image] = []
            if imageWiseAllLatencies_1st.get(ap.image) is None:
                imageWiseAllLatencies_1st[ap.image] = []
            for pe in ap.poke_events:
                if pe.latency is not None:
                    outProxy.append([pe.imageAppearanceTime, contrastLevel, pe.latency])
                    allLatencies.append(pe.latency)
                    trueLatencies.append(pe.latency)
                    rewardTimes.append(pe.imageAppearanceTime)
                    imageWiseTrueLatencies[ap.image].append(pe.latency)
                    imageWiseAllLatencies[ap.image].append(pe.latency)
                    if ap.rewardSeqNum == 1:
                        imageWiseTrueLatencies_1st[ap.image].append(pe.latency)
                        imageWiseAllLatencies_1st[ap.image].append(pe.latency)
                # NOTE that a poke event has a LATENCY of NONE iff the poke was unsuccessful.
                # Because latencies are considered only for reward images and REWARD images are
                # reset once the first poke ceases, such a case is not encountered unless
                # an erroneous wheel rotation causes event switching and falsely creates two events
                # one successful and the other unsuccessful.

    rewardImgs = set(filter(lambda im: im.imageType is ImageTypes.REWARD,
                            [ap.image for ap in Image.appearanceLog.values()]))
    for ri in rewardImgs:

        ri.true_latencies = imageWiseTrueLatencies.get(ri)
        ri.all_latencies = imageWiseAllLatencies.get(ri)

        # length of all latencies should be equal to numAppearances
        # but discrepancy may exist owing to unsuccessful pokes

        ri.true_latencies_1st = imageWiseTrueLatencies_1st.get(ri)
        ri.all_latencies_1st = imageWiseAllLatencies_1st.get(ri)

        if ri.true_latencies:
            ri.true_avg_latency = np.mean(imageWiseTrueLatencies.get(ri))
            ri.true_SEM_latency = stats.sem(imageWiseTrueLatencies.get(ri))
            ri.true_SD_latency = np.std(imageWiseTrueLatencies.get(ri))
        else:
            ri.true_avg_latency = 'N/A'
            ri.true_SEM_latency = 'N/A'
            ri.true_SD_latency = 'N/A'

        if ri.true_latencies_1st:
            ri.true_avg_latency_1st = np.mean(imageWiseTrueLatencies_1st.get(ri))
            ri.true_SEM_latency_1st = stats.sem(imageWiseTrueLatencies_1st.get(ri))
            ri.true_SD_latency_1st = np.std(imageWiseTrueLatencies_1st.get(ri))
        else:
            ri.true_avg_latency_1st = 'N/A'
            ri.true_SEM_latency_1st = 'N/A'
            ri.true_SD_latency_1st = 'N/A'

        if ri.all_latencies:
            ri.all_avg_latency = np.mean(imageWiseAllLatencies.get(ri))
            ri.all_SEM_latency = stats.sem(imageWiseAllLatencies.get(ri))
            ri.all_SD_latency = np.std(imageWiseAllLatencies.get(ri))
        else:
            ri.all_avg_latency = 'N/A'
            ri.all_SEM_latency = 'N/A'
            ri.all_SD_latency = 'N/A'

        if ri.all_latencies_1st:
            ri.all_avg_latency_1st = np.mean(imageWiseAllLatencies_1st.get(ri))
            ri.all_SEM_latency_1st = stats.sem(imageWiseAllLatencies_1st.get(ri))
            ri.all_SD_latency_1st = np.std(imageWiseAllLatencies_1st.get(ri))
        else:
            ri.all_avg_latency_1st = 'N/A'
            ri.all_SEM_latency_1st = 'N/A'
            ri.all_SD_latency_1st = 'N/A'

    outputCSV.append([])
    pokeStatistics(rewardImgs, outputCSV, preset)

    for line in outProxy:
        outputCSV.append(line)  # send latency documentation to output

    if len(trueLatencies) == 0:
        return

    ws2 = wb.create_sheet(title='All')
    headings = ["Time", "Latency", ""]
    sheetData = [[r / (60 ** 2) for r in rewardTimes], allLatencies, []]  # time in hours
    ws2.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws2.append(row)
        except ValueError:
            pass

    sheetData = []
    headings = []
    ws3 = wb.create_sheet(title='Image-wise')
    for im in sorted(imageWiseTrueLatencies.keys(), key=getContrast):
        headings.extend(["Image Contrast", "Latency"])
        headings.append("")
        latencies = imageWiseTrueLatencies.get(im)
        sheetData.append([getContrast(im)] * len(latencies) + ["", "MEAN", "SEM", "STD DEV"])
        sheetData.append(latencies + ["", im.true_avg_latency, im.true_SEM_latency, im.true_SD_latency])
        sheetData.append([])
    ws3.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws3.append(row)
        except ValueError:
            pass

    sheetData = []
    headings = []
    ws4 = wb.create_sheet(title='Distributions')
    for im in sorted(imageWiseTrueLatencies.keys(), key=getContrast):
        headings.extend(["Contrast", "Bin", "Count", "Rel. Frequency", ""])
        latencies = imageWiseTrueLatencies.get(im)
        count, hbin = np.histogram(latencies, bins=np.arange(0, TIMEOUTS.get(preset, 10) + LATENCYSTEP, LATENCYSTEP))
        count = list(count)
        hbin = list(hbin)
        sheetData.append([getContrast(im)] * len(hbin))
        sheetData.append(hbin + ["", "Total"])
        sheetData.append(count + ["", "", sum(count)])
        percents = [c * 100 / len(latencies) for c in count]
        sheetData.append(percents + ["", "", sum(percents)])  # relative frequencies as %
        sheetData.append([])
        # sheetData = list(map(lambda arr: list(map(lambda k: str(k), arr)), sheetData))
    ws4.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws4.append(row)
        except ValueError:
            pass


def pokesPerHour(poke_events, outputCSV):
    hourlyPokes = {}  # dictionary stores pokes for each hour
    for pe in poke_events:
        for t, s in zip(pe.pumpTimes, pe.pumpStates):
            if s is PumpStates.On:
                hr = int(t / 3600) + 1  # convert t to hours, round up for nth hour
                # increment pokes for each hour, default value of 0 supplied to initialize
                hourlyPokes[hr] = hourlyPokes.get(hr, 0) + 1
    outputCSV.append(['Hour', '# Successful Pokes'])
    for k in range(1, 13):
        print("Successful pokes in hour #{0} >> {1}".format(k, hourlyPokes.get(k, 0)))
        outputCSV.append([k, hourlyPokes.get(k, 0)])


# def drinkLengths(poke_events):
#     lengths = []
#     for pe in poke_events:
#         lengths.extend(pe.drinkTimes())
#     plt.hist(lengths)
#     plt.xlabel("Time (sec) drinking sugar water")
#     plt.ylabel("Frequency")
#     plt.show()


def endRun(wheelHalfTimes, image, rotation_intervals):
    if len(wheelHalfTimes) < 3:
        return
    rotation_intervals.append(RotationInterval(wheelHalfTimes, image))  # add this interval to list


def endPoke(doorStates, doorTimes, pumpTimes, pumpStates, image, poke_events):
    poke_events.append(PokeEvent(doorStates, doorTimes, pumpTimes, pumpStates, image))


def pruneRotationIntervals(rotation_intervals):
    erratic = []
    for ri in rotation_intervals:
        if not ri.viable:
            erratic.append(ri)
    for ri in erratic:
        rotation_intervals.remove(ri)


def pokeStatistics(images, outputCSV, preset):
    rewardImgs = list(filter(lambda im: im.imageType is ImageTypes.REWARD, images))

    # sort images bycontrast level
    def tryint(x):
        try:
            return int(x)
        except ValueError:
            return x

    rewardImgs.sort(key=lambda ri: [tryint(c) for c in re.split('([0-9]+)', ri.name)])
    imagePerformance(rewardImgs, outputCSV, preset)
    if preset is Presets.SPATIAL or preset is Presets.CONTRAST:
        imagePerformanceFirst(rewardImgs, outputCSV, preset)
    print('\n')


def imagePerformance(rewardImgs, outputCSV, preset):
    if preset is Presets.SPATIAL or preset is Presets.CONTRAST:
        outputCSV.append(["Image Name", "Contrast", "Appearances", "Hits", "Misses", "Success Rate %",
                          "Hits Latency Mean", "Hits Latency SEM", "Hits Latency SD", "Hits RI", "", "All Latency Mean",
                          "All Latency SEM", "All Latency SD", "All RI"])
    else:
        outputCSV.append(["Image Name", "Appearances", "Hits", "Misses", "Success Rate %",
                          "Hits Latency Mean", "Hits Latency SEM", "Hits Latency SD", "", "All Latency Mean",
                          "All Latency SEM", "All Latency SD"])
    zero_cont_mean = None
    for ri in rewardImgs:
        hits = len(ri.true_latencies) if ri.true_latencies else 0
        numAppearances = len(ri.all_latencies)

        print('REWARD image appearances for {0} >> {1}'.format(ri.name, numAppearances))
        print('Hits/Successful Pokes >> ', hits)
        success_rate = hits * 100.0 / numAppearances if numAppearances else 'N/A'

        contrast = getContrast(ri)

        if preset is Presets.SPATIAL or preset is Presets.CONTRAST:
            if zero_cont_mean is None and contrast != 0:
                raise ValueError('reward images out of order, need 0 contrast')
            elif contrast == 0:
                zero_cont_mean = ri.true_avg_latency, ri.all_avg_latency

            isNumber = lambda x: isinstance(x, numbers.Number)

            outputCSV.append([ri.name, contrast, numAppearances, hits, numAppearances - hits,
                              success_rate, ri.true_avg_latency, ri.true_SEM_latency, ri.true_SD_latency,
                              1 - zero_cont_mean[0] / ri.true_avg_latency if isNumber(zero_cont_mean[0]) and
                              isNumber(ri.true_avg_latency) else "N/A", "", ri.all_avg_latency,
                              ri.all_SEM_latency, ri.all_SD_latency,
                              1 - zero_cont_mean[1] / ri.all_avg_latency if isNumber(ri.all_avg_latency) and
                                                                            isNumber(zero_cont_mean[1]) else "N/A"])

        else:
            outputCSV.append([ri.name, numAppearances, hits, numAppearances - hits,
                              success_rate, ri.true_avg_latency, ri.true_SEM_latency, ri.true_SD_latency, "",
                              ri.all_avg_latency, ri.all_SEM_latency, ri.all_SD_latency])


def imagePerformanceFirst(rewardImgs, outputCSV, preset):
    outputCSV.append([])
    outputCSV.append(["FIRST APPEARANCES ONLY"])
    if preset is Presets.SPATIAL or preset is Presets.CONTRAST:
        outputCSV.append(["Image Name", "Contrast", "Appearances", "Hits", "Misses", "Success Rate %",
                          "Hits Latency Mean", "Hits Latency SEM", "Hits Latency SD", "Hits RI", "", "All Latency Mean",
                          "All Latency SEM", "All Latency SD", "All RI"])
    else:
        outputCSV.append(["Image Name", "Appearances", "Hits", "Misses", "Success Rate %",
                          "Hits Latency Mean", "Hits Latency SEM", "Hits Latency SD", "", "All Latency Mean",
                          "All Latency SEM", "All Latency SD"])
    zero_cont_mean = None
    for ri in rewardImgs:
        hits = len(ri.true_latencies_1st) if ri.true_latencies_1st else 0
        firstAppearances = len(ri.all_latencies_1st)
        print('FIRST ONLY REWARD image appearances for {0} >> {1}'.format(ri.name, firstAppearances))
        print('Hits/Successful Pokes >> ', hits)
        success_rate = hits * 100.0 / firstAppearances if firstAppearances else 'N/A'

        contrast = getContrast(ri)

        if preset is Presets.SPATIAL or preset is Presets.CONTRAST:
            if zero_cont_mean is None and contrast != 0:
                raise ValueError('reward images out of order, need 0 contrast')
            elif contrast == 0:
                zero_cont_mean = ri.true_avg_latency, ri.all_avg_latency

            isNumber = lambda x: isinstance(x, numbers.Number)

            outputCSV.append([ri.name, getContrast(ri), firstAppearances, hits, firstAppearances - hits,
                              success_rate, ri.true_avg_latency_1st, ri.true_SEM_latency_1st, ri.true_SD_latency_1st,
                              1 - zero_cont_mean[0] / ri.true_avg_latency_1st if isNumber(ri.true_avg_latency_1st) and
                              isNumber(zero_cont_mean[0]) else "N/A",
                              "", ri.all_avg_latency_1st, ri.all_SEM_latency_1st, ri.all_SD_latency_1st,
                              1 - zero_cont_mean[1] / ri.all_avg_latency_1st if isNumber(ri.all_avg_latency_1st) and
                              isNumber(zero_cont_mean[1]) else "N/A"])
        else:
            outputCSV.append([ri.name, firstAppearances, hits, firstAppearances - hits,
                              success_rate, ri.true_avg_latency, ri.true_SEM_latency, ri.true_SD_latency, "",
                              ri.all_avg_latency, ri.all_SEM_latency, ri.all_SD_latency])


def analyzeRotations(rotation_intervals, wb):
    #imageWiseRPMs
    rot_ints_byImage = {}
    for im, g in groupby(rotation_intervals, key=lambda ri: ri.image):
        if rot_ints_byImage.get(im, None) is None:
            rot_ints_byImage[im] = []
        rot_ints_byImage[im].extend(list(g))

    ws = wb.create_sheet(title='RPMs')
    headings = []
    sheetData = []

    for im in sorted(rot_ints_byImage.keys(), key=getContrast):
        headings.extend(["Image Contrast", "RPM", "Time"])
        headings.append("")
        rpms = rot_ints_byImage.get(im)
        sheetData.append([getContrast(im)] * len(rpms) + ["", "MEAN", "SEM", "STD DEV"])
        speedAvgs = [ri.avgSpeed for ri in rot_ints_byImage.get(im)]
        sheetData.append(speedAvgs + ["", np.mean(speedAvgs), stats.sem(speedAvgs), np.std(speedAvgs)])
        sheetData.append([ri.startTime for ri in rot_ints_byImage.get(im)])
        sheetData.append([])

    try:
        sheetData.append(["", "Global Mean RPM", "Global RPM SEM", "Global RPM STD"])
        globalAvgs = [ri.avgSpeed for ri in rotation_intervals]
        sheetData.append(["", np.mean(globalAvgs), stats.sem(globalAvgs), np.std(globalAvgs)])
    except ValueError:
        pass  # this error can only be encountered if there are no rotation intervals

    ws.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws.append(row)
        except ValueError:
            pass


def getFileNames(location):
    fileNames = []

    def recursiveDirectories(loc):
        nonlocal fileNames
        try:
            for d in next(os.walk(loc))[1]:
                recursiveDirectories(loc + d + '/')
            for f in next(os.walk(loc))[2]:
                if 'Results' in f and '.txt' in f:
                    fileNames.append(loc + f)
        except StopIteration:
            pass

    recursiveDirectories(location)
    fileNames.sort()
    return fileNames


def initialize(allInput, filename, findFloat):
    images = []
    preset = ''
    mouseNum = 0
    for line in allInput:
        if 'USB drive ID: ' in line:
            print("\n***********************************\n")
            print(filename)
            print(line)
            mouseNum = int(findFloat.search(line).group(0))
        elif 'Control image set:' in line:
            for img in line[line.find('[') + 1:line.rfind(']')].split(','):
                images.append(Image(img.strip(), ImageTypes.CONTROL))
        elif 'Reward image set:' in line:
            for img in line[line.find('[') + 1:line.rfind(']')].split(','):
                images.append(Image(img.strip(), ImageTypes.REWARD))
        elif 'preset: ' in line:
            preset = line.split('preset: ')[1]
            if 'contrast' in preset.lower():
                preset = Presets.CONTRAST
            elif 'spatial' in preset.lower():
                preset = Presets.SPATIAL
            elif '1' in preset:
                preset = Presets.NIGHT_1
            elif '2' in preset:
                preset = Presets.NIGHT_2
            elif '3' in preset:
                preset = Presets.NIGHT_3
            elif '4' in preset:
                preset = Presets.NIGHT_4

        elif "Start of experiment" in line:
            return images, "Mouse_{0}".format(mouseNum), preset


def analyze():
    global LOCALDIR
    if not LOCALDIR.endswith('/'):
        LOCALDIR += '/'
    for filename in getFileNames(LOCALDIR):
        Image.appearanceLog = OrderedDict()  # reset appearances
        with open(filename, 'r') as resultFile:
            allInput = resultFile.readlines()
        findFloat = re.compile("[+-]?([0-9]*[.])?[0-9]+")  # regex to search for a number (float)
        wheelHalfTimes, doorStates, doorTimes, pumpStates, pumpTimes, poke_events, rotation_intervals = [], [], [], [], [], [], []
        skipLine = False
        curImgName = None
        pokeInProgress = False
        try:
            images, identifier, preset = initialize(allInput, filename, findFloat)
        except TypeError as e:
            print(e.__traceback__)
            continue
        images = set(images)  # convert to set to avoid accidental duplication
        Image.images = images

        wb = Workbook()  # open(filename.replace(filename[filename.rfind('/') + 1:], identifier + '.csv'), 'w')
        outputCSV = wb.active
        try:
            controlImgStart = [im for im in images if im.imageType == ImageTypes.CONTROL][0]
        except IndexError:
            print("Warning: No CONTROL Images")
            outputCSV.append(["WARNING: no CONTROL images defined"])
            controlImgStart = [im for im in images][0]
        # ControlImgStart defined in case wheel or door activity is documented prior to first image appearance
        # documentation. This occurs rarely and is a bug in the results file generation protocol.

        currentImg, pokeImg, runImg, currentState = controlImgStart, controlImgStart, controlImgStart, None

        for line in allInput:

            if 'starting' in line:
                continue

            elif 'Image' in line and 'Name:' in line:
                newImgName = line[line.find('Name:') + 5: line.find(',')].strip()
                if curImgName != newImgName:  # ignore if it is the same image (this is a bug)
                    newImg = next((img for img in images if img.name == newImgName), None)
                    assert newImg is not None, 'Unrecognized image: {0}'.format(newImgName)
                    newImg.incrementAppearances(float(re.search("Time: (.*)", line).group(1)), currentImg)
                    curImgName = newImgName
                    currentImg = newImg


            elif 'Wheel' in line and not pokeInProgress:
                if skipLine:
                    skipLine = False
                    continue
                if currentState is Activity.Poking:
                    endPoke(doorStates, doorTimes, pumpTimes, pumpStates, pokeImg, poke_events)
                    doorStates, doorTimes, pumpTimes, pumpStates = [], [], [], []
                currentState = Activity.Running
                if 'State:' in line:
                    wheelHalfTimes.append(float(findFloat.search(line).group(0)))  # appends times
                if 'revolution' in line:
                    # need to skip next data point because wheel state does not actually change; it appears to be a bug
                    skipLine = True
                    continue  # do NOT reset skipLine boolean

            elif 'Pump' in line:
                if re.search("State: (.*), Time", line).group(1) == 'On':
                    pump_state = PumpStates.On
                    pokeImg = currentImg  # the poke event's image should be the image when the pump is on (ie REWARD image)
                    pokeInProgress = True  # ensure parameters don't change within poke duration
                else:
                    pump_state = PumpStates.Off
                    pokeInProgress = False
                pumpStates.append(pump_state)
                pumpTimes.append(float(findFloat.search(line).group(0)))

            elif 'Door' in line:
                if currentState is Activity.Running:
                    endRun(wheelHalfTimes, currentImg, rotation_intervals)
                    wheelHalfTimes = []
                if currentState is not Activity.Poking and not pokeInProgress:
                    pokeImg = currentImg  # record image when poke event starts
                currentState = Activity.Poking
                door_state = DoorStates.High if re.search("State: (.*), Time", line).group(
                    1) == 'High' else DoorStates.Low
                doorStates.append(door_state)
                doorTimes.append(float(findFloat.search(line).group(0)))

            skipLine = False
        if currentState is Activity.Poking:
            endPoke(doorStates, doorTimes, pumpTimes, pumpStates, pokeImg, poke_events)
        else:
            endRun(wheelHalfTimes, currentImg, rotation_intervals)
        pruneRotationIntervals(rotation_intervals)

        analysisFuncs(poke_events, rotation_intervals, wb, preset)
        wb.save(filename.replace(filename[filename.rfind('/') + 1:], identifier + '.xlsx'))


'''ANALYSIS FUNCTION CALLS BEGIN HERE; DO NOT EDIT ABOVE WHEN RUNNING ANALYSIS. CHANGES SHOULD BE MADE ONLY TO 
analysisFuncs METHOD BELOW.'''


def analysisFuncs(poke_events, rotation_intervals, wb, preset):
    outputCSV = wb.active
    pokeLatencies(wb, preset)
    pokesPerHour(poke_events, outputCSV)  # Note that 'outputCSV' is the first sheet in the workbook 'wb'.
    analyzeRotations(rotation_intervals, wb)


analyze()
